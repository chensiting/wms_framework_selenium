# -*- coding: utf-8 -*-
# @File : 926keywords.py
# @Time : 2024-01-28 20:48
# @Author : chensiting
# @Email : 475707078@qq.com
# @Software: PyCharm


import pdfplumber
import re
import openpyxl

ad=[]
k=[]
for l in range(1,19):
    adress=f"D:/下载/926词汇笔记/926逐词精讲-词汇{l}笔记.pdf"
    ad.append(adress)
    pdf = pdfplumber.open(adress) #将导入的pdf打开
    print(f'正在写入{adress}')
    for i in range(len(pdf.pages)): #获取pdf总页数，并遍历
        page = pdf.pages[i]  #打开pdf对应页数
        wk = page.extract_text()  #获取对应页数的文本内容
        k.append(wk)  #将内容压入列表
wb = ''.join(k)   #将列表的内容遍历拼接并转为字符串类型
# #将pdf的全部内容压入一个列表后，通过.join函数将列表遍历拼接转为字符串类型，为接下来正则匹配做准备。

wb = wb.replace('改变就橙啦橙啦，大学生学习成长平台'," ")
num = re.findall('(\d+).*?\[', wb)
text = re.findall('\d+(.*?)\[', wb)
chinese = re.findall('](.*?)【真题例句',wb,re.S)   #re.S的作用是使匹配包含换行符\n、\t，即空格与换行

xs = openpyxl.Workbook()
xl = xs.active
xl.title = '达叔926'
xl['A1']='序号'
xl['B1']='英文'
xl['C1']='中文|助记'
q=2
for a,b,c in zip(num,text,chinese):
    xl.cell(q,1,a)
    xl.cell(q,2,b)
    xl.cell(q,3,c)
    q+=1
xs.save('达叔926.xlsx')
